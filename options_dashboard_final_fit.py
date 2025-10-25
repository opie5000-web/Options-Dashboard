# options_dashboard_final_fit.py
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
import numpy as np
from streamlit_autorefresh import st_autorefresh
import time

# ----------------- CONFIG -----------------
CHART_HEIGHT = 300
TABLE_HEIGHT = 350
COMPACT_MODE = True

# Auto-refresh every 30 seconds
st_autorefresh(interval=30000)

st.set_page_config(
    page_title="Live Options Data Dashboard",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- COMPACT + ALIGNMENT CSS ---
st.markdown("""
<style>
.stApp { background-color: #000; }

/* Start just below Streamlit gray bar */
.block-container {
    padding-top: 0.6rem !important;
    margin-top: 0rem !important;
}

/* Tight first block */
div[data-testid="stVerticalBlock"] > div:first-child {
    margin-top: -0.2rem !important;
}

/* Heading + layout tweaks */
h3 {
    margin-top: 0rem !important;
    padding-top: 0rem !important;
}

/* Reduce space between elements globally */
div[data-testid="stVerticalBlock"] div[data-testid="stVerticalBlock"] {
    margin-top: 0rem !important;
    padding-top: 0rem !important;
}

/* Remove padding Streamlit adds between elements */
.block-container > div {
    padding-top: 0rem !important;
    margin-top: 0rem !important;
}

/* Table + metric styling */
.stDataFrame td, .stDataFrame th {
    font-size: 13px !important;
    color: #fff !important;
    background: #000 !important;
}
.stMetric > label { 
    font-size: 8px !important; 
    color: #fff !important; 
    margin-bottom: 0 !important; 
}
.stMetric .stMetricValue { 
    font-size: 14px !important; 
}
.stMetric { 
    margin: 0 !important; 
    padding: 0 !important; 
    height: auto !important;
}
.element-container .stPlotlyChart { padding: 0.25rem; }

/* Progress bar slimmer */
.stProgress { 
    height: 6px !important; 
    margin: 2px 0 0 0 !important; 
}
.stProgress > div > div { 
    height: 4px !important; 
}

/* Reduce space above/below radio buttons */
div[data-testid="stRadio"] {
    margin-top: -0.25rem !important;
    margin-bottom: -0.25rem !important;
}
</style>
""", unsafe_allow_html=True)

# ----------------- DATA LOADER -----------------
@st.cache_data(ttl=30)
def load_data():
    wb = load_workbook('NewData.xlsx', read_only=True, data_only=True)
    ws = wb['ChartData']

    strikes, gex_oi, abs_oi, gex_vol, abs_vol = [], [], [], [], []
    r = 2
    while r <= ws.max_row:
        s = ws.cell(row=r, column=1).value
        if s is None:
            break
        strikes.append(s)
        gex_oi.append(float(ws.cell(row=r, column=2).value or 0))
        abs_oi.append(float(ws.cell(row=r, column=3).value or 0))
        gex_vol.append(float(ws.cell(row=r, column=4).value or 0))
        abs_vol.append(float(ws.cell(row=r, column=5).value or 0))
        r += 1

    spot = float(ws.cell(row=2, column=8).value or 0)
    gex_oi_gauge = float(ws.cell(row=2, column=15).value or 0) * 100
    gex_vol_gauge = float(ws.cell(row=2, column=17).value or 0) * 100

    table_data = []
    for row_num in range(2, 14):
        label = ws.cell(row=row_num, column=7).value
        if label is None or 'ABS' in str(label):
            continue
        qqq_val = ws.cell(row=row_num, column=8).value
        nq_val = ws.cell(row=row_num, column=9).value
        qqq = float(qqq_val or 0)
        nq = round(float(nq_val or 0) * 4) / 4
        table_data.append({'Label': label, 'QQQ': qqq, 'NQ': nq})
    df_table = pd.DataFrame(table_data)

    min_len = min(len(strikes), len(gex_oi), len(abs_oi), len(gex_vol), len(abs_vol))
    strikes = strikes[:min_len]
    gex_oi = gex_oi[:min_len]
    abs_oi = abs_oi[:min_len]
    gex_vol = gex_vol[:min_len]
    abs_vol = abs_vol[:min_len]

    pos_gex_oi = [max(0, x) for x in gex_oi]
    neg_gex_oi = [min(0, x) for x in gex_oi]
    pos_gex_vol = [max(0, x) for x in gex_vol]
    neg_gex_vol = [min(0, x) for x in gex_vol]

    return (strikes, pos_gex_oi, neg_gex_oi, abs_oi,
            pos_gex_vol, neg_gex_vol, abs_vol,
            spot, gex_oi_gauge, gex_vol_gauge, df_table)

# Load data
(strikes, pos_gex_oi, neg_gex_oi, abs_oi,
 pos_gex_vol, neg_gex_vol, abs_vol,
 spot, gex_oi_gauge, gex_vol_gauge, df_table) = load_data()

# ----------------- LAYOUT -----------------
col1, col2, col3 = st.columns([2, 2, 1])

with col1:
    st.markdown("<h4 style='color:white;margin-bottom:0.3rem; font-size:14px;'>GEX-OI Chart</h4>", unsafe_allow_html=True)
    fig_oi = make_subplots(specs=[[{"secondary_y": True}]])
    max_pos_idx = np.argmax(pos_gex_oi) if any(pos_gex_oi) else -1
    max_neg_idx = np.argmax([-x for x in neg_gex_oi]) if any(neg_gex_oi) else -1
    pos_colors = ['#00ff00' if i == max_pos_idx else '#007bff' for i in range(len(pos_gex_oi))]
    neg_colors = ['#ff0000' if i == max_neg_idx else '#007bff' for i in range(len(neg_gex_oi))]
    fig_oi.add_trace(go.Bar(x=strikes, y=pos_gex_oi, name='Positive', marker_color=pos_colors, width=0.8))
    fig_oi.add_trace(go.Bar(x=strikes, y=neg_gex_oi, name='Negative', marker_color=neg_colors, width=0.8))
    fig_oi.add_trace(go.Scatter(x=strikes, y=abs_oi, name='ABS-OI', line=dict(color='#6f42c1', width=2)), secondary_y=True)

    fig_oi.add_vline(x=spot, line_dash="dash", line_color="white", line_width=0.5)

    fig_oi.update_layout(
        autosize=True,
        height=CHART_HEIGHT,
        plot_bgcolor='black',
        paper_bgcolor='black',
        font=dict(color='white', size=11),
        bargap=0.05,
        bargroupgap=0.1,
        margin=dict(l=12, r=8, t=6, b=6),
        showlegend=False
    )
    fig_oi.update_xaxes(showticklabels=True, showgrid=False)
    fig_oi.update_yaxes(showticklabels=False, showgrid=False)
    st.plotly_chart(fig_oi, use_container_width=True)

    progress_value = min(1.0, abs(gex_oi_gauge) / 100.0)
    bar_color = '#00ff00' if gex_oi_gauge >= 0 else '#ff0000'
    st.markdown(f"""
    <div style="background-color: #333; height: 6px; border-radius: 3px; margin-top: 4px;">
      <div style="background-color: {bar_color}; height: 100%; width: {progress_value*100}%; border-radius: 3px;"></div>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.markdown("<h4 style='color:white;margin-bottom:0.3rem; font-size:14px;'>GEX-VOL Chart</h4>", unsafe_allow_html=True)
    fig_vol = make_subplots(specs=[[{"secondary_y": True}]])
    max_pos_idx = np.argmax(pos_gex_vol) if any(pos_gex_vol) else -1
    max_neg_idx = np.argmax([-x for x in neg_gex_vol]) if any(neg_gex_vol) else -1
    pos_colors = ['#00ff00' if i == max_pos_idx else '#007bff' for i in range(len(pos_gex_vol))]
    neg_colors = ['#ff0000' if i == max_neg_idx else '#007bff' for i in range(len(neg_gex_vol))]
    fig_vol.add_trace(go.Bar(x=strikes, y=pos_gex_vol, name='Positive', marker_color=pos_colors, width=0.8))
    fig_vol.add_trace(go.Bar(x=strikes, y=neg_gex_vol, name='Negative', marker_color=neg_colors, width=0.8))
    fig_vol.add_trace(go.Scatter(x=strikes, y=abs_vol, name='ABS-VOL', line=dict(color='#8b5cf6', width=2)), secondary_y=True)

    fig_vol.add_vline(x=spot, line_dash="dash", line_color="white", line_width=0.5)

    fig_vol.update_layout(
        autosize=True,
        height=CHART_HEIGHT,
        plot_bgcolor='black',
        paper_bgcolor='black',
        font=dict(color='white', size=11),
        bargap=0.05,
        bargroupgap=0.1,
        margin=dict(l=12, r=8, t=6, b=6),
        showlegend=False
    )
    fig_vol.update_xaxes(showticklabels=True, showgrid=False)
    fig_vol.update_yaxes(showticklabels=False, showgrid=False)
    st.plotly_chart(fig_vol, use_container_width=True)

    progress_value = min(1.0, abs(gex_vol_gauge) / 100.0)
    bar_color = '#00ff00' if gex_vol_gauge >= 0 else '#ff0000'
    st.markdown(f"""
    <div style="background-color: #333; height: 6px; border-radius: 3px; margin-top: 4px;">
      <div style="background-color: {bar_color}; height: 100%; width: {progress_value*100}%; border-radius: 3px;"></div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    st.markdown("<div style='color:white; font-weight:600; margin-bottom:4px;'>Summary</div>", unsafe_allow_html=True)

    def color_row(row): return ['background-color: #000000; color: white'] * len(row)
    styled_df = df_table.style.apply(color_row, axis=1).format({'QQQ': '{:.1f}', 'NQ': '{:.1f}'})
    st.dataframe(styled_df, height=TABLE_HEIGHT, use_container_width=True, hide_index=True)

# Footer
st.markdown(
    f"<div style='color:#999; font-size:12px; margin-top:6px;'>Last updated: {time.strftime('%Y-%m-%d %H:%M:%S')} | Auto-refreshing every 30s</div>",
    unsafe_allow_html=True,
)