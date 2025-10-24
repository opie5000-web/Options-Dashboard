import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
import numpy as np
from streamlit_autorefresh import st_autorefresh
import time

# Auto-refresh every 30 seconds
st_autorefresh(interval=30000)

# Set page config for better layout
st.set_page_config(
    page_title="Live Options Data Dashboard",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("---")

# Load data (use read-only mode for live Excel)
@st.cache_data(ttl=30)
def load_data():
    wb = load_workbook('NewData.xlsx', read_only=True, data_only=True)
    ws_chart = wb['ChartData']
    
    # Extract chart data (assuming headers in row 1, data from row 2; drop NaNs)
    strikes = []
    gex_oi = []
    abs_oi = []
    gex_vol = []
    abs_vol = []
    
    row = 2
    while row <= ws_chart.max_row:
        strike_cell = ws_chart.cell(row=row, column=1).value  # Col A: Strike
        if strike_cell is None:
            break
        strikes.append(strike_cell)
        
        gex_oi.append(float(ws_chart.cell(row=row, column=2).value or 0))  # Col B: GEX-OI
        abs_oi.append(float(ws_chart.cell(row=row, column=3).value or 0))  # Col C: ABS-OI
        gex_vol.append(float(ws_chart.cell(row=row, column=4).value or 0))  # Col D: GEX-VOL
        abs_vol.append(float(ws_chart.cell(row=row, column=5).value or 0))  # Col E: ABS-VOL
        
        row += 1
    
    spot = float(ws_chart.cell(row=2, column=8).value or 0)  # Col H, first data row
    gex_oi_gauge = float(ws_chart.cell(row=2, column=15).value or 0) * 100  # Col O, first data row
    gex_vol_gauge = float(ws_chart.cell(row=2, column=17).value or 0) * 100  # Col Q, first data row
    
    # Table data (cols G:I, all rows but exclude ABS labels)
    table_data = []
    for row_num in range(2, 14):  # Original range for 12 rows
        label = ws_chart.cell(row=row_num, column=7).value  # Col G
        if label is None or 'ABS' in str(label):  # Skip ABS rows
            continue
        qqq_val = ws_chart.cell(row=row_num, column=8).value  # Col H
        nq_val = ws_chart.cell(row=row_num, column=9).value  # Col I
        qqq = float(qqq_val or 0)
        nq = round(float(nq_val or 0) * 4) / 4
        table_data.append({'Label': label, 'QQQ': qqq, 'NQ': nq})
    df_table = pd.DataFrame(table_data)
    
    # Trim lists to same length
    min_len = min(len(strikes), len(gex_oi), len(abs_oi), len(gex_vol), len(abs_vol))
    strikes = strikes[:min_len]
    gex_oi = gex_oi[:min_len]
    abs_oi = abs_oi[:min_len]
    gex_vol = gex_vol[:min_len]
    abs_vol = abs_vol[:min_len]
    
    # Positive/negative splits
    pos_gex_oi = [max(0, x) for x in gex_oi]
    neg_gex_oi = [min(0, x) for x in gex_oi]
    pos_gex_vol = [max(0, x) for x in gex_vol]
    neg_gex_vol = [min(0, x) for x in gex_vol]
    
    return (strikes, pos_gex_oi, neg_gex_oi, abs_oi, pos_gex_vol, neg_gex_vol, abs_vol, 
            spot, gex_oi_gauge, gex_vol_gauge, df_table)

# Load data
strikes, pos_gex_oi, neg_gex_oi, abs_oi, pos_gex_vol, neg_gex_vol, abs_vol, spot, gex_oi_gauge, gex_vol_gauge, df_table = load_data()

# Main layout: Wider columns for better balance [4:1]
col1, col2 = st.columns([4, 1])

with col1:
    st.subheader("GEX Options Chart")
    
    # Move radio button above the chart for better flow
    view_type = st.radio("View:", ["GEX-VOL", "GEX-OI"], horizontal=True, key="chart_view")
    
    # Create subplots for dual y-axes
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    
    if view_type == "GEX-OI":
        fig.add_trace(go.Bar(x=strikes, y=pos_gex_oi, name='GEX-OI Positive', 
                             marker_color='#007bff'), secondary_y=False)
        fig.add_trace(go.Bar(x=strikes, y=neg_gex_oi, name='GEX-OI Negative', 
                             marker_color='#dc3545'), secondary_y=False)
        fig.add_trace(go.Scatter(x=strikes, y=abs_oi, name='ABS-OI', 
                                 line=dict(color='#6f42c1', width=2)), secondary_y=True)
    else:
        fig.add_trace(go.Bar(x=strikes, y=pos_gex_vol, name='GEX-VOL Positive', 
                             marker_color='#007bff'), secondary_y=False)
        fig.add_trace(go.Bar(x=strikes, y=neg_gex_vol, name='GEX-VOL Negative', 
                             marker_color='#dc3545'), secondary_y=False)
        fig.add_trace(go.Scatter(x=strikes, y=abs_vol, name='ABS-VOL', 
                                 line=dict(color='#8b5cf6', width=2)), secondary_y=True)
    
    # Spot line
    fig.add_vline(x=spot, line_dash="dash", line_color="white", 
                  annotation_text=f"Spot {spot}", annotation_position="top")
    
    # Improved layout - Remove all gridlines
    fig.update_layout(
        height=600,  # Taller for better visibility
        title_text="Options Data",
        plot_bgcolor='black',
        paper_bgcolor='black',
        font=dict(color='white', size=12),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.2,
            xanchor="center",
            x=0.5,
            bgcolor='rgba(0,0,0,0)'
        ),
        xaxis=dict(
            showgrid=False,  # Remove vertical grids
            tickmode='linear',
            dtick=1,  # Show every strike price (assuming integers)
            tickfont=dict(color='white'),
            title="Strike Price"
        ),
        yaxis=dict(
            showgrid=False,  # Remove horizontal grids
            tickfont=dict(color='white'),
            title="GEX (Stacked)"
        ),
        yaxis2=dict(
            showgrid=False,  # Remove horizontal grids on secondary y
            tickfont=dict(color='white'),
            title="ABS",
            overlaying='y',
            side='right'
        )
    )
    st.plotly_chart(fig, use_container_width=True)

with col2:
    # Add some top padding
    st.markdown("###")
    
    # No subheader for Summary Table
    
    # Improved table styling and sizing - Sized for 10 rows
    def color_row(row):
        label = row['Label']
        if label in ['PG-OI', 'FG-OI', 'PG-TT', 'FG-TT']:
            return ['background-color: #00ff00; color: black'] * len(row)
        elif label in ['FR-OI', 'NG-OI', 'FR-TT', 'NG-TT']:
            return ['background-color: #ff0000; color: white'] * len(row)
        return [None] * len(row)
    
    styled_df = df_table.style.apply(color_row, axis=1).format({'QQQ': '{:.1f}', 'NQ': '{:.1f}'})
    st.dataframe(
        styled_df, 
        use_container_width=True, 
        height=350,  # Slightly taller to fit all rows without scrolling
        hide_index=True
    )
    
    st.markdown("---")  # Separator
    
    # No subheader for GEX Gauge
    
    # Gauge radio inside section
    gauge_type = st.radio("Gauge Type:", ["GEX-VOL", "GEX-OI"], horizontal=True, key="gauge_type")
    gauge_val = gex_vol_gauge if gauge_type == "GEX-VOL" else gex_oi_gauge
    
    # Better gauge visualization: Use a simple bar-like metric with color
    col_g1, col_g2 = st.columns(2)
    with col_g1:
        st.metric("Gauge Value", f"{gauge_val:.1f}%")
    with col_g2:
        gauge_color = "🟢" if gauge_val >= 0 else "🔴"
        st.metric("Status", gauge_color)
    
    # Progress bar with better styling
    progress_val = abs(gauge_val) / 100.0
    progress_bar = st.progress(min(1.0, progress_val))
    
    # Add direction indicator
    direction = "Positive" if gauge_val >= 0 else "Negative"
    st.caption(f"Direction: {direction}")

# Footer caption, full width
st.markdown("---")
col_full = st.columns(1)
with col_full[0]:
    st.caption(f"Last updated: {time.strftime('%Y-%m-%d %H:%M:%S')} | Auto-refreshing every 30s")

# Optional: Custom CSS for dark theme enhancements and larger table font
st.markdown("""
<style>
    .stApp {
        background-color: black;
    }
    .stMetric > label {
        color: white;
    }
    .stMetric > div > div > div {
        color: white;
    }
    .stDataFrame {
        background-color: #111;
    }
    .stDataFrame td, .stDataFrame th {
        font-size: 16px !important;
    }
</style>
""", unsafe_allow_html=True)